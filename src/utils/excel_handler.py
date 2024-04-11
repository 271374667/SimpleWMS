from dataclasses import dataclass, fields
from pathlib import Path
from typing import Dict, Optional, Sequence, Type, Union

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler:
    """Excel处理类,用于导入导出excel
    Methods
        get_data: 获取数据
        get_header: 获取表头
        get_header_type: 获取表头类型
        set_dataclass: 设置数据类
        set_data: 设置数据
        export2excel: 导出到excel
        read_from_excel: 从excel读取
    """

    def __init__(self):
        self._current_dataclass: Optional[Type[dataclass]] = None
        self._header_type: Optional[Dict[str, Type]] = None
        self._header: Optional[list[str]] = None
        self._data: Optional[Sequence[dataclass]] = None

    def get_data(self) -> Sequence[dataclass]:
        return self._data

    def get_header(self) -> list[str]:
        return self._header

    def get_header_type(self) -> Dict[str, Type]:
        return self._header_type

    def set_dataclass(self, data: Type[dataclass]) -> None:
        self._current_dataclass = data
        self._header_type = {field.name: field.type for field in fields(data)}
        self._header = list(self._header_type.keys())

    def set_data(self, data: Sequence[dataclass]) -> None:
        self._data = data

    def export2excel(self, output_path: Union[Path, str]) -> None:

        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active
        ws.append(self._header)
        for item in self._data:
            ws.append([getattr(item, field) for field in self._header])
        wb.save(output_path)

    def read_from_excel(self, input_path: Path) -> None:
        wb = openpyxl.load_workbook(input_path)
        ws: Worksheet = wb.active
        self._data = []
        self._data.extend(
            self._current_dataclass(*row)
            for row in list(ws.iter_rows(values_only=True))[1:]
        )


if __name__ == "__main__":
    from dataclasses import dataclass
    from datetime import date
    from src.core.constant import ROOT_DIR
    import tempfile

    @dataclass
    class Person:
        name: str
        age: int
        birthday: date

    data = [
        Person("张三", 18, date(2000, 1, 1)),
        Person("李四", 20, date(1998, 1, 1)),
    ]

    excel_handler = ExcelHandler()
    excel_handler.set_dataclass(Person)
    excel_handler.set_data(data)

    out_file = ROOT_DIR / "test.xlsx"
    # excel_handler.export2excel(out_file)
    # excel_handler.read_from_excel(out_file)
    # print(excel_handler.get_data())
    with tempfile.TemporaryDirectory() as tmpdirname:
        excel_path = Path(tmpdirname) / "test.xlsx"
        excel_handler.export2excel(excel_path)
        excel_handler.read_from_excel(excel_path)
        print(excel_handler.get_data())
        print(excel_handler.get_header())
        print(excel_handler.get_header_type())

    # Output:
    # [Person(name='张三', age=18, birthday=datetime.date(2000, 1, 1)), Person(name='李四', age=20, birthday=datetime.date(1998, 1, 1))]
    # ['name', 'age', 'birthday']
    # {'name': <class 'str'>, 'age': <class 'int'>, 'birthday': <class 'datetime.date'>}
